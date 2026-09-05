# utils.py
import streamlit as st
import pandas as pd
import numpy as np
import os
import re
import unicodedata
import json
import sqlite3
import bcrypt
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from typing import Dict, Optional, List, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import requests
import time
from pathlib import Path

# =============================================
# CONSTANTES DE CAMINHOS (CSVs)
# =============================================
ARQUIVO_CSV_PROFISSIONAL = "perfil_completo_jogadores_profissional_2026.csv"
ARQUIVO_CSV_SUB15 = "perfil_completo_jogadores_Sub15_2026.csv"
ARQUIVO_CSV_SUB17 = "perfil_completo_jogadores_Sub17_2026.csv"
ARQUIVO_CSV_COMISSAO_PROFISSIONAL = "perfil_completo_comissao_2026.csv"
ARQUIVO_CSV_COMISSAO_SUB15 = "perfil_completo_comissao_Sub15_2026.csv"
ARQUIVO_CSV_COMISSAO_SUB17 = "perfil_completo_comissao_Sub17_2026.csv"
ARQUIVO_LESOES_PROFISSIONAL = "jogadores_linhares_profissional_lesoes.csv"
ARQUIVO_LESOES_SUB15 = "jogadores_linhares_Sub15_lesoes.csv"
ARQUIVO_LESOES_SUB17 = "jogadores_linhares_Sub17_lesoes.csv"
ARQUIVO_BIO_PROFISSIONAL = "jogadores_linhares_profissional_Bioimpedancia.csv"
ARQUIVO_BIO_SUB15 = "jogadores_linhares_Sub15_Bioimpedancia.csv"
ARQUIVO_BIO_SUB17 = "jogadores_linhares_Sub17_Bioimpedancia.csv"
ARQUIVO_CRONO_PROF = "cronograma_profissional_2026.csv"
ARQUIVO_CRONO_SUB15 = "cronograma_sub15_2026.csv"
ARQUIVO_CRONO_SUB17 = "cronograma_sub17_2026.csv"

DATA_DIR = "data"
RELATORIOS_DIR = "relatorios"
NOME_TIME = "Linhares FC"
TEMPORADA = str(datetime.now().year)

# =============================================
# CONSTANTES DE PASTAS DE ESTATÍSTICAS
# =============================================
PASTA_ESTATISTICAS_PROFISSIONAL = "data/estatisticas_jogadores/"
PASTA_ESTATISTICAS_SUB15 = "data/estatisticas_sub15/"
PASTA_ESTATISTICAS_SUB17 = "data/estatisticas_sub17/"
PASTA_ESTATISTICAS_COMISSAO_PROFISSIONAL = "data/estatisticas_comissao_tecnica_profissional/"
PASTA_ESTATISTICAS_COMISSAO_SUB15 = "data/estatisticas_comissao_tecnica_sub15/"
PASTA_ESTATISTICAS_COMISSAO_SUB17 = "data/estatisticas_comissao_tecnica_sub17/"

# =============================================
# CONSTANTES DOS ARQUIVOS DE CARTÕES (JSON)
# =============================================
CAMINHO_CARTOES_PROFISSIONAL = "cartoes_acumulados_profissional.json"
CAMINHO_CARTOES_SUB15 = "cartoes_acumulados_sub15.json"
CAMINHO_CARTOES_SUB17 = "cartoes_acumulados_sub17.json"
CAMINHO_CARTOES_COMISSAO_PROFISSIONAL = "cartoes_acumulados_comissao_profissional.json"
CAMINHO_CARTOES_COMISSAO_SUB15 = "cartoes_acumulados_comissao_sub15.json"
CAMINHO_CARTOES_COMISSAO_SUB17 = "cartoes_acumulados_comissao_sub17.json"

# =============================================
# CONFIGURAÇÕES POR CATEGORIA
# =============================================
CATEGORIA_CONFIG = {
    "Profissional": {
        "team_id": 12928,
        "competicao_id": 2,
        "elenco_func": "carregar_elenco_profissional",
        "cartoes_key": "profissional",
    },
    "Sub-15": {
        "team_id": 27831,
        "competicao_id": 11,
        "elenco_func": "carregar_elenco_sub15",
        "cartoes_key": "sub15",
    },
    "Sub-17": {
        "team_id": 27832,
        "competicao_id": 10,
        "elenco_func": "carregar_elenco_sub17",
        "cartoes_key": "sub17",
    }
}

# =============================================
# MAPEAMENTO DE NOMES (JOGADORES E COMISSÃO)
# =============================================
MAPEAMENTO_NOMES_PROFISSIONAL = {
    'Wenderson Silva Neves': 'Wendy',
    'Wendy': 'Wendy',
    'Marcus Paulo Sousa Oliveira': 'Marcus Paulo',
    'Marcus Paulo': 'Marcus Paulo',
    'Francisco Wesley da Silva Sousa': 'Wesley',
    'Wesley': 'Wesley',
    'Francisco de Assis Rapozo Neto': 'Francisco Neto',
    'Francisco Neto': 'Francisco Neto',
    'Stuart Asafe Ferreira Alves': 'Stuart',
    'Stuart': 'Stuart',
    'Yuri Ribeiro Giovanelli': 'Yuri Ribeiro',
    'Yuri Ribeiro': 'Yuri Ribeiro',
    'João Pedro Firmino Oliveira': 'João Firmino',
    'João Firmino': 'João Firmino',
    'Joao Firmino': 'João Firmino',
    'Lucas Titol Lopes': 'Lucas Titol',
    'Lucas Titol': 'Lucas Titol',
    'Rayner Silva Gomes': 'Rayner',
    'Rayner': 'Rayner',
    'Kayque Santos da Cunha': 'Kayque Santos',
    'Kayque Santos': 'Kayque Santos',
    'Cayque': 'Kayque Santos',
    'Ruan Amaral Rios': 'Ruan Rios',
    'Ruan Rios': 'Ruan Rios',
    'Genilson dos Santos Júnior': 'Júnior Espeto',
    'Júnior Espeto': 'Junior Espeto',
    'Clavis Severo Leão': 'Clavis Neto',
    'Clavis Neto': 'Clavis Neto',
    'Jeferson David Palacios Cantillo': 'Jeferson Palacios',
    'Jeferson Palacios': 'Jeferson Palacios',
    'J. D. Palacios Cantillo': 'Jeferson Palacios',
    'Gabriel Amorim de Aguiar': 'Gabriel Amorim',
    'Gabriel Amorim': 'Gabriel Amorim',
    'Virgílio Santos Borges': 'Borjão',
    'Borjão': 'Borjão',
    'Borjao': 'Borjão',
    'Matheus Toribes Ferreira Souza': 'Matheus Toribes',
    'Matheus Toribes': 'Matheus Toribes',
    'João Marcos Santos Ferraz Luz': 'João Marcos',
    'João Marcos': 'João Marcos',
    'Davi Fornaciari Lima': 'Davi Fornaciari',
    'Davi Fornaciari': 'Davi Fornaciari',
    'Karlos Henrique dos Reis Calavort': 'Kaká',
    'Kaká': 'Kaká',
    'Kaka': 'Kaká',
    'Daniel Olmo Morais Gonçalves': 'Daniel Olmo',
    'Daniel Olmo': 'Daniel Olmo',
    'Arthur Luiz Darros': 'Arthur Darros',
    'Arthur Darros': 'Arthur Darros',
    'Júlio César Fontana Leite': 'Julio César',
    'Julio César': 'Julio César',
    'Matheus Sarmento Mesquita': 'Matheus Sarmento',
    'Matheus Sarmento': 'Matheus Nossa',
    'Gabriel de Jesus Rodrigues': 'Gabriel Jesus',
    'Gabriel Jesus': 'Gabriel Jesus',
}
MAPEAMENTO_NOMES_SUB15 = {}
MAPEAMENTO_NOMES_SUB17 = {}
MAPEAMENTO_NOMES_COMISSAO_PROFISSIONAL = {}
MAPEAMENTO_NOMES_COMISSAO_SUB15 = {}
MAPEAMENTO_NOMES_COMISSAO_SUB17 = {}

# =============================================
# ATRIBUTOS FM26 (GLOBAL)
# =============================================
ATRIBUTOS_FM26 = [
    'escanteios', 'cruzamentos', 'drible', 'finalizacao', 'primeiro_controle',
    'cobranca_faltas', 'cabecada', 'chutes_longe', 'arremessos_laterais',
    'marcacao', 'passe', 'cobranca_penaltis', 'desarme', 'tecnica',
    'agressividade', 'antecipacao', 'coragem', 'composicao', 'concentracao',
    'decisao', 'determinacao', 'criatividade', 'lideranca', 'movimentacao_sem_bola',
    'posicionamento', 'trabalho_equipe', 'visao_jogo', 'intensidade_trabalho',
    'aceleracao', 'agilidade', 'equilibrio', 'altura_salto', 'condicao_fisica_natural',
    'velocidade_maxima', 'resistencia', 'forca_fisica', 'reflexos', 'jogo_aereo_goleiro',
    'defesas_goleiro', 'comando_area', 'comunicacao_goleiro', 'chutes_goleiro',
    'um_contra_um_goleiro', 'saida_gol', 'tendencia_socar', 'arremessos_goleiro',
    'excentricidade', 'consistencia', 'jogo_sujo', 'jogos_importantes',
    'propensao_lesao', 'versatilidade', 'adaptabilidade', 'ambicao', 'lealdade',
    'pressao', 'profissionalismo', 'esportividade', 'temperamento', 'controversia'
]

# =============================================
# FUNÇÃO PARA SANITIZAR DATAFRAMES (EVITA ERRO DE ARROW)
# =============================================
def sanitizar_dataframe(df):
    if df is None or df.empty:
        return df
    df = df.copy()
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].apply(
                lambda x: ', '.join(x) if isinstance(x, list) else (str(x) if pd.notna(x) else '')
            )
        elif pd.api.types.is_categorical_dtype(df[col]):
            df[col] = df[col].astype(str)
        elif pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime('%Y-%m-%d')
    return df

# =============================================
# FUNÇÃO PARA ORDENAR HISTÓRICO DE CARTÕES
# =============================================
def ordenar_historico_cartoes(cartoes: dict) -> dict:
    for jogador, dados in cartoes.items():
        if 'historico' in dados and dados['historico']:
            dados['historico'] = sorted(
                dados['historico'],
                key=lambda x: datetime.strptime(x['data'], "%Y-%m-%d") if x['data'] and '/' not in x['data'] else datetime.strptime(x['data'], "%d/%m/%Y") if '/' in x['data'] else datetime.now()
            )
    return cartoes

# =============================================
# FUNÇÕES AUXILIARES BÁSICAS
# =============================================
def calcular_idade(data_nasc_str, data_referencia=None):
    if pd.isna(data_nasc_str) or not data_nasc_str:
        return np.nan
    try:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                data_nasc = datetime.strptime(str(data_nasc_str).strip(), fmt)
                break
            except ValueError:
                continue
        else:
            return np.nan
        hoje = data_referencia if data_referencia else datetime.now()
        if isinstance(hoje, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                try:
                    hoje = datetime.strptime(hoje, fmt)
                    break
                except ValueError:
                    continue
            if isinstance(hoje, str):
                return np.nan
        idade = hoje.year - data_nasc.year - ((hoje.month, hoje.day) < (data_nasc.month, data_nasc.day))
        return idade
    except Exception:
        return np.nan

def normalizar_texto(texto):
    if pd.isna(texto):
        return ""
    texto = str(texto).strip().lower()
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    texto = re.sub(r'[^a-z0-9\s]', '', texto)
    return ' '.join(texto.split())

def mapear_nome_para_canonico(nome):
    if pd.isna(nome):
        return None
    nome = str(nome).strip()
    if nome in MAPEAMENTO_NOMES_PROFISSIONAL:
        return MAPEAMENTO_NOMES_PROFISSIONAL[nome]
    if nome in MAPEAMENTO_NOMES_SUB15:
        return MAPEAMENTO_NOMES_SUB15[nome]
    if nome in MAPEAMENTO_NOMES_SUB17:
        return MAPEAMENTO_NOMES_SUB17[nome]
    if nome in MAPEAMENTO_NOMES_COMISSAO_PROFISSIONAL:
        return MAPEAMENTO_NOMES_COMISSAO_PROFISSIONAL[nome]
    if nome in MAPEAMENTO_NOMES_COMISSAO_SUB15:
        return MAPEAMENTO_NOMES_COMISSAO_SUB15[nome]
    if nome in MAPEAMENTO_NOMES_COMISSAO_SUB17:
        return MAPEAMENTO_NOMES_COMISSAO_SUB17[nome]
    nome_norm = normalizar_texto(nome)
    for var, can in MAPEAMENTO_NOMES_PROFISSIONAL.items():
        if normalizar_texto(var) == nome_norm:
            return can
    for var, can in MAPEAMENTO_NOMES_SUB15.items():
        if normalizar_texto(var) == nome_norm:
            return can
    for var, can in MAPEAMENTO_NOMES_SUB17.items():
        if normalizar_texto(var) == nome_norm:
            return can
    for var, can in MAPEAMENTO_NOMES_COMISSAO_PROFISSIONAL.items():
        if normalizar_texto(var) == nome_norm:
            return can
    for var, can in MAPEAMENTO_NOMES_COMISSAO_SUB15.items():
        if normalizar_texto(var) == nome_norm:
            return can
    for var, can in MAPEAMENTO_NOMES_COMISSAO_SUB17.items():
        if normalizar_texto(var) == nome_norm:
            return can
    return nome

def safe_str(valor, padrao="N/I"):
    if pd.isna(valor) or str(valor).strip() == '':
        return padrao
    return str(valor).strip()

def extrair_id_jogo(caminho_arquivo):
    nome = os.path.basename(caminho_arquivo)
    match = re.search(r'jogo_(\d+)_', nome)
    if match:
        return int(match.group(1))
    return None

def extrair_data_jogo(caminho_arquivo):
    match = re.search(r'(\d{4}-\d{2}-\d{2})', caminho_arquivo)
    if match:
        try:
            return datetime.strptime(match.group(1), "%Y-%m-%d")
        except:
            pass
    return None

# =============================================
# CLASSIFICAÇÕES (IMC, GORDURA, ESTADO FÍSICO)
# =============================================
def classif_imc(imc):
    if pd.isna(imc): return "Indefinido"
    if imc < 20: return "Baixo peso"
    if imc < 24: return "Normal"
    if imc < 27: return "Sobrepeso leve"
    if imc < 30: return "Sobrepeso"
    return "Obesidade"

def classif_gordura(p, idade):
    if pd.isna(p) or pd.isna(idade): return "Indefinido"
    if idade < 30:
        if p < 12: return "Excelente"
        if p < 17: return "Bom"
        if p < 22: return "Médio"
        return "Alto"
    else:
        if p < 15: return "Excelente"
        if p < 20: return "Bom"
        if p < 25: return "Médio"
        return "Alto"

def estado_fisico(imc_class, gor_class):
    if imc_class == "Indefinido" or gor_class == "Indefinido": return "Bom"
    if imc_class == "Normal" and gor_class in ["Excelente", "Bom"]: return "Ótimo"
    if imc_class == "Normal" and gor_class == "Médio": return "Bom"
    if imc_class in ["Sobrepeso leve", "Sobrepeso"]: return "Atenção"
    if imc_class == "Obesidade" or gor_class == "Alto": return "Crítico"
    return "Regular"

# =============================================
# INICIALIZAÇÃO DO BANCO SQLITE
# =============================================
def inicializar_banco():
    conn = sqlite3.connect('meu_futebol.db')
    cursor = conn.cursor()
    # Criação das tabelas (já existentes, mantidas)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS treinos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            atleta_id TEXT,
            data TEXT,
            carga REAL,
            duracao_min INTEGER
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS wellbeing (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            atleta_id TEXT,
            data TEXT,
            sono INTEGER,
            estresse INTEGER,
            dor INTEGER,
            disposicao INTEGER
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS lesoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            jogador TEXT,
            tipo_lesao TEXT,
            data_inicio TEXT,
            data_fim TEXT,
            ativo INTEGER DEFAULT 1
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS jogos (
            id INTEGER PRIMARY KEY,
            time_casa_id INTEGER,
            time_fora_id INTEGER,
            gols_casa INTEGER,
            gols_fora INTEGER,
            status TEXT,
            data_hora TEXT,
            formacao_casa TEXT,
            formacao_fora TEXT,
            venue_id INTEGER,
            arbitro_id INTEGER,
            competicao_id INTEGER
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS gps (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            atleta_id TEXT,
            data TEXT,
            distancia_total REAL,
            velocidade_max REAL,
            sprints INTEGER
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS times (
            id INTEGER PRIMARY KEY,
            nome TEXT,
            sigla TEXT,
            logo_url TEXT,
            fundado INTEGER,
            pais TEXT,
            temporada INTEGER,
            venue_id INTEGER,
            id_principal INTEGER
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS elenco (
            id INTEGER PRIMARY KEY,
            nome TEXT,
            apelido TEXT,
            posicao TEXT,
            numero INTEGER,
            idade INTEGER,
            foto TEXT,
            time_id INTEGER,
            data_nascimento TEXT,
            cidade_nascimento TEXT,
            uf_nascimento TEXT,
            pais_nascimento TEXT,
            competicao_id INTEGER
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS eventos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            jogo_id INTEGER,
            tempo INTEGER,
            tipo TEXT,
            jogador_id INTEGER,
            detalhes TEXT,
            time_id INTEGER,
            membro_id INTEGER,
            tipo_alvo TEXT,
            competicao_id INTEGER,
            fonte TEXT DEFAULT 'api'
        )
    ''')
    # Garante a coluna 'fonte'
    cursor.execute("PRAGMA table_info(eventos)")
    colunas = [col[1] for col in cursor.fetchall()]
    if 'fonte' not in colunas:
        cursor.execute("ALTER TABLE eventos ADD COLUMN fonte TEXT DEFAULT 'api'")
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tecnicos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT,
            cargo TEXT,
            idade INTEGER,
            data_nascimento TEXT,
            historico_profissional TEXT,
            historico_jogador TEXT,
            nacionalidade TEXT,
            foto TEXT,
            time_id INTEGER,
            cidade TEXT,
            uf TEXT,
            pais TEXT,
            competicao_id INTEGER,
            apelido TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS arbitros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            foto TEXT,
            categoria TEXT,
            uf TEXT,
            genero TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS comissao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            cargo TEXT NOT NULL,
            foto TEXT,
            data_nascimento TEXT,
            cidade TEXT,
            uf TEXT,
            pais TEXT,
            historico_profissional TEXT,
            historico_jogador TEXT,
            categoria TEXT,
            time_id INTEGER,
            apelido TEXT,
            competicao_id INTEGER,
            idade INTEGER
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS competicoes (
            id INTEGER PRIMARY KEY,
            nome TEXT NOT NULL,
            categoria TEXT,
            nivel TEXT,
            genero TEXT,
            temporada INTEGER,
            ativa BOOLEAN
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS venues (
            id INTEGER PRIMARY KEY,
            nome TEXT,
            cidade TEXT,
            capacidade INTEGER,
            endereco TEXT,
            superficie TEXT,
            imagem TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS estatisticas_jogadores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            jogo_id INTEGER,
            jogador_id INTEGER,
            minutos INTEGER,
            gols INTEGER,
            assistencias INTEGER,
            cartoes_amarelos INTEGER,
            cartoes_vermelhos INTEGER,
            chutes INTEGER,
            chutes_ao_gol INTEGER,
            desarmes INTEGER,
            interceptacoes INTEGER,
            passes_certos INTEGER,
            passes_chave INTEGER,
            defesas INTEGER,
            competicao_id INTEGER
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS lineup (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            jogo_id INTEGER NOT NULL,
            competicao_id INTEGER,
            time TEXT NOT NULL,
            time_id INTEGER,
            formacao TEXT,
            status TEXT,
            nome TEXT NOT NULL,
            numero TEXT,
            posicao TEXT,
            posicao_grid TEXT,
            jogador_id INTEGER,
            data_importacao DATETIME,
            tecnico_id INTEGER,
            comissao_id INTEGER
        )
    ''')
    conn.commit()
    conn.close()

# =============================================
# FUNÇÃO AUXILIAR PARA CARREGAR ELENCO (GENÉRICA)
# =============================================
def _carregar_elenco_generico(caminho_arquivo: str) -> pd.DataFrame:
    if not os.path.exists(caminho_arquivo):
        st.warning(f"Arquivo não encontrado: {caminho_arquivo}")
        return pd.DataFrame()
    try:
        df = pd.read_csv(caminho_arquivo, sep=';', encoding='utf-8-sig', on_bad_lines='skip')
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

        if 'id_jogo' in df.columns:
            st.warning(f"⚠️ O arquivo {caminho_arquivo} parece ser de estatísticas (coluna 'id_jogo'). Ignorando.")
            return pd.DataFrame()

        coluna_nome = None
        for possivel in ['nome_completo', 'apelido', 'jogador', 'nome']:
            if possivel in df.columns:
                coluna_nome = possivel
                break

        if coluna_nome is None:
            st.warning(f"⚠️ Nenhuma coluna de nome encontrada em {caminho_arquivo}. Ignorando.")
            return pd.DataFrame()

        if coluna_nome != 'nome_completo':
            df.rename(columns={coluna_nome: 'nome_completo'}, inplace=True)

        if 'apelido' in df.columns:
            df['nome_completo'] = df['nome_completo'].fillna(df['apelido'])
        df['nome_completo'] = df['nome_completo'].fillna('N/I')

        if 'ogol_id' in df.columns:
            df = df.drop_duplicates(subset=['ogol_id'], keep='first')
        else:
            df = df.drop_duplicates(subset=['nome_completo'], keep='first')

        for col in ['apelido', 'data_nascimento', 'posicao']:
            if col not in df.columns:
                df[col] = None

        for col in ['altura_cm', 'peso_kg', 'habilidade_atual', 'habilidade_potencial']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
            else:
                df[col] = np.nan

        for attr in ATRIBUTOS_FM26:
            if attr in df.columns:
                df[attr] = pd.to_numeric(df[attr], errors='coerce')
            else:
                df[attr] = np.nan

        df['IMC'] = df.apply(
            lambda x: x['peso_kg'] / ((x['altura_cm'] / 100) ** 2)
            if pd.notna(x['altura_cm']) and pd.notna(x['peso_kg']) and x['altura_cm'] > 0
            else np.nan,
            axis=1
        ).round(1)
        df['Classificacao_IMC'] = df['IMC'].apply(classif_imc)
        df['Idade'] = df['data_nascimento'].apply(lambda x: calcular_idade(x) if pd.notna(x) else np.nan)
        df['Gordura_Corporal_%'] = df.apply(
            lambda row: round((1.20 * row['IMC']) + (0.23 * row['Idade']) - 16.2, 1)
            if pd.notna(row['IMC']) and pd.notna(row['Idade'])
            else np.nan,
            axis=1
        )
        df['Massa_Magra_kg'] = df.apply(
            lambda row: round(row['peso_kg'] * (1 - row['Gordura_Corporal_%'] / 100), 1)
            if pd.notna(row['peso_kg']) and pd.notna(row['Gordura_Corporal_%'])
            else np.nan,
            axis=1
        )
        df['Massa_Muscular_Estimada_kg'] = df.apply(
            lambda row: round(row['Massa_Magra_kg'] * 0.55, 1)
            if pd.notna(row['Massa_Magra_kg'])
            else np.nan,
            axis=1
        )
        df['Classificacao_Gordura'] = df.apply(
            lambda x: classif_gordura(x['Gordura_Corporal_%'], x['Idade']),
            axis=1
        )
        df['Estado_Fisico'] = df.apply(
            lambda row: estado_fisico(row['Classificacao_IMC'], row['Classificacao_Gordura']),
            axis=1
        )

        def cat_pos(pos_str):
            if pd.isna(pos_str):
                return 'Outros', []
            pos = str(pos_str).upper().strip()
            pos_list = [p.strip() for p in pos.split('/')] if '/' in pos else [pos.strip()]
            cats = []
            for p in pos_list:
                pu = p.upper()
                if 'GOLEIRO' in pu:
                    cats.append('Goleiro')
                elif 'ZAGUEIRO' in pu:
                    cats.append('Zagueiro')
                elif 'LATERAL DIREITO' in pu or 'LAT. DIREITO' in pu:
                    cats.append('Lateral Direito')
                elif 'LATERAL ESQUERDO' in pu or 'LAT. ESQUERDO' in pu:
                    cats.append('Lateral Esquerdo')
                elif 'LATERAL' in pu:
                    cats.append('Lateral')
                elif 'VOLANTE' in pu:
                    cats.append('Volante')
                elif 'MEIA-CENTRAL' in pu or 'MEIA CENTRAL' in pu or 'MEIO-CENTRO' in pu:
                    cats.append('Meia-Central')
                elif 'MEIA-ATACANTE' in pu or 'MEIA ATACANTE' in pu or 'MEIA OFENSIVO' in pu:
                    cats.append('Meia-Atacante')
                elif 'MEIA' in pu or 'MEIO' in pu:
                    cats.append('Meia')
                elif 'PONTA DIREITA' in pu:
                    cats.append('Ponta Direita')
                elif 'PONTA ESQUERDA' in pu:
                    cats.append('Ponta Esquerda')
                elif 'PONTA' in pu:
                    cats.append('Ponta')
                elif 'CENTROAVANTE' in pu:
                    cats.append('Centroavante')
                elif 'SEGUNDO ATACANTE' in pu:
                    cats.append('Segundo Atacante')
                elif 'ATACANTE' in pu:
                    cats.append('Atacante')
                else:
                    cats.append('Outros')
            cats = [c for c in cats if c != 'Outros']
            cats = list(dict.fromkeys(cats))
            return cats[0] if cats else 'Outros', cats

        res = df['posicao'].apply(cat_pos)
        df['Posicao_Principal'] = res.apply(lambda x: x[0])
        df['Posicoes_Secundarias'] = res.apply(lambda x: x[1])

        df['Rating_Geral_FM26'] = df.apply(
            lambda row: min(100, row['habilidade_atual'] / 2)
            if pd.notna(row.get('habilidade_atual'))
            else 50,
            axis=1
        )

        if 'foto' in df.columns:
            df.drop(columns=['foto'], inplace=True)

        return sanitizar_dataframe(df)

    except Exception as e:
        st.error(f"❌ Erro ao carregar {caminho_arquivo}: {e}")
        return pd.DataFrame()

@st.cache_data
def carregar_elenco_profissional() -> pd.DataFrame:
    caminho = ARQUIVO_CSV_PROFISSIONAL
    if not os.path.exists(caminho):
        caminho = os.path.join(DATA_DIR, ARQUIVO_CSV_PROFISSIONAL)
    print(f"📂 Carregando profissional: {caminho}")
    return _carregar_elenco_generico(caminho)

@st.cache_data
def carregar_elenco_sub15() -> pd.DataFrame:
    caminho = ARQUIVO_CSV_SUB15
    if not os.path.exists(caminho):
        caminho = os.path.join(DATA_DIR, ARQUIVO_CSV_SUB15)
    print(f"📂 Carregando sub15: {caminho}")
    return _carregar_elenco_generico(caminho)

@st.cache_data
def carregar_elenco_sub17() -> pd.DataFrame:
    caminho = ARQUIVO_CSV_SUB17
    if not os.path.exists(caminho):
        caminho = os.path.join(DATA_DIR, ARQUIVO_CSV_SUB17)
    print(f"📂 Carregando sub17: {caminho}")
    return _carregar_elenco_generico(caminho)

# =============================================
# CARREGAMENTO DA COMISSÃO (COM TODOS OS ATRIBUTOS)
# =============================================
def _carregar_comissao_generico(caminho_arquivo: str) -> pd.DataFrame:
    if not os.path.exists(caminho_arquivo):
        return pd.DataFrame()
    try:
        df = pd.read_csv(caminho_arquivo, sep=';', encoding='utf-8-sig')
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        if 'nome' not in df.columns and 'apelido' in df.columns:
            df['nome'] = df['apelido']
        elif 'nome' not in df.columns and 'nome_completo' in df.columns:
            df['nome'] = df['nome_completo']
        if 'cargo' not in df.columns:
            df['cargo'] = 'Técnico'
        if 'idade' not in df.columns and 'data_nascimento' in df.columns:
            df['idade'] = df['data_nascimento'].apply(lambda x: calcular_idade(x) if pd.notna(x) else np.nan)
        if 'cidade_uf' not in df.columns and 'cidade_nascimento' in df.columns and 'uf_nascimento' in df.columns:
            df['cidade_uf'] = df['cidade_nascimento'].fillna('') + ', ' + df['uf_nascimento'].fillna('')
            df['cidade_uf'] = df['cidade_uf'].str.rstrip(', ')
        elif 'cidade_uf' not in df.columns:
            df['cidade_uf'] = 'N/I'
        if 'pais' not in df.columns and 'pais_nascimento' in df.columns:
            df['pais'] = df['pais_nascimento']
        elif 'pais' not in df.columns:
            df['pais'] = 'N/I'
        if 'nome_canonico' not in df.columns and 'apelido' in df.columns:
            df['nome_canonico'] = df['apelido'].apply(mapear_nome_para_canonico)
        elif 'nome_canonico' not in df.columns and 'nome' in df.columns:
            df['nome_canonico'] = df['nome'].apply(mapear_nome_para_canonico)
        return sanitizar_dataframe(df)
    except Exception as e:
        st.error(f"Erro ao carregar comissão de {caminho_arquivo}: {e}")
        return pd.DataFrame()

@st.cache_data
def carregar_comissao() -> pd.DataFrame:
    return _carregar_comissao_generico(ARQUIVO_CSV_COMISSAO_PROFISSIONAL)

@st.cache_data
def carregar_comissao_sub15() -> pd.DataFrame:
    return _carregar_comissao_generico(ARQUIVO_CSV_COMISSAO_SUB15)

@st.cache_data
def carregar_comissao_sub17() -> pd.DataFrame:
    return _carregar_comissao_generico(ARQUIVO_CSV_COMISSAO_SUB17)

# =============================================
# CRONOGRAMA
# =============================================
@st.cache_data
def carregar_cronograma(categoria="Profissional") -> pd.DataFrame:
    if categoria == "Profissional":
        arquivo = ARQUIVO_CRONO_PROF
    elif categoria == "Sub-15":
        arquivo = ARQUIVO_CRONO_SUB15
    elif categoria == "Sub-17":
        arquivo = ARQUIVO_CRONO_SUB17
    else:
        return pd.DataFrame()
    if not os.path.exists(arquivo):
        return pd.DataFrame()
    try:
        df = pd.read_csv(arquivo, sep=';', encoding='utf-8-sig')
        if 'data' not in df.columns:
            return pd.DataFrame()
        df['data'] = pd.to_datetime(df['data'], errors='coerce')
        if 'competicao' not in df.columns:
            df['competicao'] = 'Desconhecida'
        else:
            df['competicao'] = df['competicao'].fillna('Desconhecida')
        if 'fase' not in df.columns:
            df['fase'] = ''
        else:
            df['fase'] = df['fase'].fillna('')
        return sanitizar_dataframe(df)
    except Exception as e:
        st.error(f"Erro ao carregar cronograma: {e}")
        return pd.DataFrame()

def obter_proximo_jogo(categoria="Profissional") -> Optional[Dict]:
    df = carregar_cronograma(categoria)
    if df.empty:
        return None
    hoje = datetime.now().date()
    df_futuros = df[df['data'].dt.date >= hoje].sort_values('data')
    if df_futuros.empty:
        return None
    return df_futuros.iloc[0].to_dict()

# =============================================
# EXIBIR FOTO (JOGADORES E COMISSÃO)
# =============================================
def obter_caminho_foto(pessoa_row, categoria="Profissional"):
    if 'foto' in pessoa_row.index:
        foto = pessoa_row.get('foto')
        if foto and pd.notna(foto) and str(foto).strip():
            caminho = str(foto).strip()
            if os.path.exists(caminho) or caminho.startswith('http'):
                return caminho

    nome = pessoa_row.get('apelido') or pessoa_row.get('nome_completo') or pessoa_row.get('nome')
    if nome:
        nome_clean = normalizar_texto(nome).replace(' ', '_')
        pastas = [
            "fotos_sistema_Analise_Elenco/Jogadores/Profissional",
            "fotos_sistema_Analise_Elenco/Jogadores/Sub15",
            "fotos_sistema_Analise_Elenco/Jogadores/Sub17",
            "fotos_sistema_Analise_Elenco/Comissao_Tecnica/Profissional",
            "fotos_sistema_Analise_Elenco/Comissao_Tecnica/Sub15",
            "fotos_sistema_Analise_Elenco/Comissao_Tecnica/Sub17",
            "fotos"
        ]
        for ext in ['.png', '.jpg', '.jpeg']:
            for pasta in pastas:
                caminho = os.path.join(pasta, f"{nome}{ext}")
                if os.path.exists(caminho):
                    return caminho
                caminho = os.path.join(pasta, f"{nome_clean}{ext}")
                if os.path.exists(caminho):
                    return caminho
    return None

def exibir_foto(pessoa_row, categoria="Profissional", width=100):
    caminho = obter_caminho_foto(pessoa_row, categoria)
    if caminho and (caminho.startswith('http') or os.path.exists(caminho)):
        try:
            st.image(caminho, width=width)
            return
        except:
            pass
    st.write("📷")

# =============================================
# FOTO DE ÁRBITROS
# =============================================
def obter_caminho_foto_arbitro(nome_arbitro: str) -> Optional[str]:
    if not nome_arbitro or pd.isna(nome_arbitro):
        return None

    base_dir = Path(__file__).resolve().parent
    nome_clean = normalizar_texto(nome_arbitro).replace(' ', '_')
    extensoes = ['.png', '.jpg', '.jpeg', '.webp']

    pastas = [
        base_dir / "fotos_arbitros",
        base_dir / "assets" / "fotos_arbitros",
        base_dir / "data" / "fotos_arbitros",
    ]

    for pasta in pastas:
        if not pasta.exists():
            continue
        for ext in extensoes:
            caminho = pasta / f"{nome_arbitro}{ext}"
            if caminho.exists():
                return str(caminho.resolve())
            caminho = pasta / f"{nome_clean}{ext}"
            if caminho.exists():
                return str(caminho.resolve())

    for pasta in pastas:
        if pasta.exists():
            for ext in extensoes:
                for arquivo in pasta.rglob(f"*{ext}"):
                    if arquivo.stem.lower() == nome_clean.lower() or arquivo.stem == nome_arbitro:
                        return str(arquivo.resolve())
    return None

# =============================================
# LESÕES (CSV)
# =============================================
def parse_data_flexivel(data_str):
    if pd.isna(data_str) or str(data_str).strip() == '':
        return None
    data_str = str(data_str).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(data_str, fmt).date()
        except ValueError:
            continue
    return None

def formatar_data_br(data_date):
    return data_date.strftime("%d/%m/%Y") if data_date else ''

def carregar_lesoes(categoria):
    if categoria == 'profissional':
        csv_path = ARQUIVO_LESOES_PROFISSIONAL
    elif categoria == 'sub15':
        csv_path = ARQUIVO_LESOES_SUB15
    elif categoria == 'sub17':
        csv_path = ARQUIVO_LESOES_SUB17
    else:
        return {}, {}
    if not csv_path or not os.path.exists(csv_path):
        return {}, {}
    try:
        df = pd.read_csv(csv_path, delimiter=';', encoding='utf-8-sig', dtype=str)
    except Exception:
        return {}, {}
    lesionados_por_ogol = {}
    lesionados_por_nome = {}
    colunas_lesoes = [col for col in df.columns if col.startswith('Lesao_')]
    for idx, row in df.iterrows():
        ogol_id = row.get('ogol_id')
        nome = row.get('nome_completo')
        lesionado = False
        for col in colunas_lesoes:
            valor = row.get(col, '')
            if pd.notna(valor) and valor != '':
                ocorrencias = str(valor).split(',')
                ultima = ocorrencias[-1].strip()
                if '/' in ultima or '-' in ultima:
                    data_obj = parse_data_flexivel(ultima)
                    if data_obj is not None:
                        if '/' not in ultima:
                            lesionado = True
                            break
        if ogol_id and pd.notna(ogol_id):
            try:
                ogol_id_int = int(float(ogol_id))
                lesionados_por_ogol[ogol_id_int] = lesionado
            except:
                pass
        if nome:
            lesionados_por_nome[nome] = lesionado
    return lesionados_por_ogol, lesionados_por_nome

def adicionar_coluna_lesionado(df, categoria):
    les_ogol, les_nome = carregar_lesoes(categoria)
    def is_lesionado(row):
        ogol = row.get('ogol_id')
        if pd.notna(ogol):
            try:
                ogol_int = int(float(ogol))
                if ogol_int in les_ogol:
                    return les_ogol[ogol_int]
            except:
                pass
        nome = row.get('nome_completo')
        if nome and nome in les_nome:
            return les_nome[nome]
        return False
    df['lesionado'] = df.apply(is_lesionado, axis=1)
    return df

def obter_historico_lesoes_texto(jogador_row, categoria):
    if categoria == 'Profissional':
        csv_path = ARQUIVO_LESOES_PROFISSIONAL
    elif categoria == 'Sub-15':
        csv_path = ARQUIVO_LESOES_SUB15
    elif categoria == 'Sub-17':
        csv_path = ARQUIVO_LESOES_SUB17
    else:
        return "Categoria inválida."
    if not csv_path or not os.path.exists(csv_path):
        return "Arquivo de lesões não encontrado."
    try:
        df_lesoes = pd.read_csv(csv_path, delimiter=';', encoding='utf-8-sig', dtype=str)
    except Exception as e:
        return f"Erro ao ler lesões: {e}"
    ogol_id = jogador_row.get('ogol_id')
    nome = jogador_row.get('nome_completo')
    linha_lesao = None
    if pd.notna(ogol_id):
        try:
            ogol_int = int(float(ogol_id))
            linha_lesao = df_lesoes[df_lesoes['ogol_id'].astype(float).astype(int) == ogol_int]
        except:
            pass
    if linha_lesao is None or linha_lesao.empty:
        linha_lesao = df_lesoes[df_lesoes['nome_completo'] == nome]
    if linha_lesao is None or linha_lesao.empty:
        return "Nenhum registro de lesão encontrado."
    colunas_lesoes = [col for col in df_lesoes.columns if col.startswith('Lesao_')]
    if not colunas_lesoes:
        return "Nenhuma coluna de lesão definida no CSV."
    linhas = []
    tem_lesao = False
    for col in colunas_lesoes:
        valor = linha_lesao.iloc[0].get(col, '')
        if pd.notna(valor) and valor != '':
            tem_lesao = True
            nome_lesao = col.replace('Lesao_', '').replace('_', ' ')
            ocorrencias = str(valor).split(',')
            ocorrencias_formatadas = []
            for occ in ocorrencias:
                occ = occ.strip()
                if '/' in occ:
                    if ' - ' in occ:
                        data_inicio_str, data_fim_str = occ.split(' - ', 1)
                    else:
                        data_inicio_str, data_fim_str = occ.split('/', 1)
                    data_inicio = parse_data_flexivel(data_inicio_str.strip())
                    data_fim = parse_data_flexivel(data_fim_str.strip())
                    if data_inicio and data_fim:
                        ocorrencias_formatadas.append(f"{formatar_data_br(data_inicio)} - {formatar_data_br(data_fim)}")
                    else:
                        ocorrencias_formatadas.append(occ)
                else:
                    data_obj = parse_data_flexivel(occ)
                    if data_obj:
                        ocorrencias_formatadas.append(f"{formatar_data_br(data_obj)} (atual)")
                    else:
                        ocorrencias_formatadas.append(occ)
            linhas.append(f"• {nome_lesao}: {', '.join(ocorrencias_formatadas)}")
    if not tem_lesao:
        return "Nenhuma lesão registrada."
    return "\n".join(linhas)

def obter_lesao_atual(jogador_row, categoria):
    if categoria == 'Profissional':
        csv_path = ARQUIVO_LESOES_PROFISSIONAL
    elif categoria == 'Sub-15':
        csv_path = ARQUIVO_LESOES_SUB15
    elif categoria == 'Sub-17':
        csv_path = ARQUIVO_LESOES_SUB17
    else:
        return ""
    if not csv_path or not os.path.exists(csv_path):
        return ""
    try:
        df_lesoes = pd.read_csv(csv_path, delimiter=';', encoding='utf-8-sig', dtype=str)
    except Exception:
        return ""
    ogol_id = jogador_row.get('ogol_id')
    nome = jogador_row.get('nome_completo')
    linha_lesao = None
    if pd.notna(ogol_id):
        try:
            ogol_int = int(float(ogol_id))
            linha_lesao = df_lesoes[df_lesoes['ogol_id'].astype(float).astype(int) == ogol_int]
        except:
            pass
    if linha_lesao is None or linha_lesao.empty:
        linha_lesao = df_lesoes[df_lesoes['nome_completo'] == nome]
    if linha_lesao is None or linha_lesao.empty:
        return ""
    colunas_lesoes = [col for col in df_lesoes.columns if col.startswith('Lesao_')]
    for col in colunas_lesoes:
        valor = linha_lesao.iloc[0].get(col, '')
        if pd.notna(valor) and str(valor).strip() != '':
            ocorrencias = str(valor).split(',')
            ultima = ocorrencias[-1].strip()
            if ' - ' in ultima or '/' in ultima or '–' in ultima:
                continue
            nome_lesao = col.replace('Lesao_', '').replace('_', ' ')
            return nome_lesao
    return ""

# =============================================
# GESTÃO DE LESÕES (ADICIONAR/ATUALIZAR)
# =============================================
def adicionar_lesao(csv_path, nome_jogador, tipo_lesao, data_inicio, data_fim=None):
    import pandas as pd
    if not os.path.exists(csv_path):
        df = pd.DataFrame(columns=['ogol_id', 'nome_completo'])
    else:
        try:
            df = pd.read_csv(csv_path, delimiter=';', encoding='utf-8-sig', dtype=str,
                             on_bad_lines='skip')
        except Exception as e:
            print(f"Erro ao ler {csv_path}: {e}")
            return

    if 'nome_completo' not in df.columns:
        df['nome_completo'] = ''
    if 'ogol_id' not in df.columns:
        df['ogol_id'] = ''

    if nome_jogador in df['nome_completo'].values:
        idx = df[df['nome_completo'] == nome_jogador].index[0]
    else:
        nova_linha = {'nome_completo': nome_jogador, 'ogol_id': ''}
        for col in df.columns:
            if col not in nova_linha:
                nova_linha[col] = ''
        df = pd.concat([df, pd.DataFrame([nova_linha])], ignore_index=True)
        idx = len(df) - 1

    df.at[idx, 'nome_completo'] = nome_jogador

    coluna_lesao = f"Lesao_{tipo_lesao.replace(' ', '_').title()}"
    if coluna_lesao not in df.columns:
        df[coluna_lesao] = ''

    valor_atual = df.at[idx, coluna_lesao]
    if pd.isna(valor_atual) or valor_atual == '':
        nova_ocorrencia = data_inicio if data_fim is None else f"{data_inicio} - {data_fim}"
    else:
        if data_fim is None:
            nova_ocorrencia = f"{valor_atual}, {data_inicio}"
        else:
            nova_ocorrencia = f"{valor_atual}, {data_inicio} - {data_fim}"

    df.at[idx, coluna_lesao] = nova_ocorrencia
    df.to_csv(csv_path, sep=';', encoding='utf-8-sig', index=False)

def adicionar_lesao_com_data_fim(csv_path, nome_jogador, tipo_lesao, data_fim):
    import pandas as pd
    if not os.path.exists(csv_path):
        return

    try:
        df = pd.read_csv(csv_path, delimiter=';', encoding='utf-8-sig', dtype=str,
                         on_bad_lines='skip')
    except Exception as e:
        print(f"Erro ao ler {csv_path}: {e}")
        return

    coluna_lesao = f"Lesao_{tipo_lesao.replace(' ', '_').title()}"
    if coluna_lesao not in df.columns:
        return

    if nome_jogador in df['nome_completo'].values:
        idx = df[df['nome_completo'] == nome_jogador].index[0]
    else:
        return

    valor = df.at[idx, coluna_lesao]
    if pd.isna(valor) or valor == '':
        return

    ocorrencias = str(valor).split(',')
    for i in range(len(ocorrencias) - 1, -1, -1):
        occ = ocorrencias[i].strip()
        if ' - ' not in occ:
            ocorrencias[i] = f"{occ} - {data_fim}"
            break

    df.at[idx, coluna_lesao] = ', '.join(ocorrencias)
    df.to_csv(csv_path, sep=';', encoding='utf-8-sig', index=False)

# =============================================
# BIOIMPEDÂNCIA (CSV)
# =============================================
def carregar_dados_bioimpedancia(categoria):
    if categoria == 'profissional':
        csv_path = ARQUIVO_BIO_PROFISSIONAL
    elif categoria == 'sub15':
        csv_path = ARQUIVO_BIO_SUB15
    elif categoria == 'sub17':
        csv_path = ARQUIVO_BIO_SUB17
    else:
        return {}
    if not csv_path or not os.path.exists(csv_path):
        return {}
    try:
        df = pd.read_csv(csv_path, delimiter=';', encoding='utf-8-sig', dtype=str)
    except Exception:
        return {}
    resultados = {}
    for idx, row in df.iterrows():
        try:
            nome = str(row.get('nome_completo', '')).strip()
            ogol_id = row.get('ogol_id')
            if pd.notna(ogol_id):
                try:
                    ogol_id = int(float(ogol_id))
                except:
                    ogol_id = None
            data_coleta = row.get('data_bioimpedancia', '')
            idade = None
            if 'data_nascimento' in row and pd.notna(row.get('data_nascimento')):
                idade = calcular_idade(row.get('data_nascimento'), data_coleta)
            altura_cm = para_float(row.get('altura_cm'))
            peso = para_float(row.get('peso_kg'))
            triceps = para_float(row.get('dobra_triceps'))
            subescap = para_float(row.get('dobra_subescapular'))
            suprail = para_float(row.get('dobra_suprailiaca'))
            abdominal = para_float(row.get('dobra_abdominal'))
            coxa = para_float(row.get('dobra_coxa'))
            panturrilha = para_float(row.get('dobra_panturilha'))
            peitoral = para_float(row.get('dobra_peitoral'))
            axiliar = para_float(row.get('dobra_axiliar_media'))
            perim_braco = para_float(row.get('perimetro_braco'))
            perim_coxa = para_float(row.get('perimetro_coxa'))
            perim_perna = para_float(row.get('perimetro_perna'))
            pct_f = None
            pct_p3 = None
            pct_p7 = None
            mm_lee = None
            if altura_cm and peso:
                imc = peso / ((altura_cm/100)**2)
                if idade:
                    pct_f = 1.20 * imc + 0.23 * idade - 16.2
                    pct_p3 = pct_f
                    pct_p7 = pct_f
                if perim_braco and perim_coxa and perim_perna and triceps and coxa and panturrilha:
                    mm_lee = 0.5 * (perim_braco + perim_coxa + perim_perna) - 0.1 * (triceps + coxa + panturrilha)
            if pct_f is not None and peso is not None:
                massa_gorda = (pct_f / 100) * peso
                massa_magra = peso - massa_gorda
            else:
                massa_gorda = None
                massa_magra = None
            dados = {
                'pct_faulkner': pct_f,
                'pct_pollock3': pct_p3,
                'pct_pollock7': pct_p7,
                'massa_gorda': massa_gorda,
                'massa_magra': massa_magra,
                'massa_muscular': mm_lee,
                'data_coleta': data_coleta,
                'peso': peso,
                'altura': altura_cm / 100.0 if altura_cm else None,
                'idade': idade
            }
            if ogol_id:
                resultados[ogol_id] = dados
            else:
                resultados[nome] = dados
        except Exception:
            continue
    return resultados

def para_float(valor):
    if pd.isna(valor) or valor == '':
        return None
    try:
        return float(valor.replace(',', '.'))
    except:
        return None

def aplicar_dados_bioimpedancia(df, dados_bio):
    if not dados_bio:
        return df
    for col in ['PctGordura_Faulkner', 'PctGordura_Pollock3', 'PctGordura_Pollock7', 'Massa_Gorda_kg']:
        if col not in df.columns:
            df[col] = np.nan
    if 'Massa_Muscular_Origem' not in df.columns:
        df['Massa_Muscular_Origem'] = ''
    for idx, row in df.iterrows():
        ogol_id = row.get('ogol_id')
        nome = row.get('nome_completo')
        bio = None
        if pd.notna(ogol_id) and ogol_id in dados_bio:
            bio = dados_bio[ogol_id]
        elif nome in dados_bio:
            bio = dados_bio[nome]
        if bio is not None:
            if bio.get('peso') is not None:
                df.at[idx, 'peso_kg'] = bio['peso']
            if bio.get('altura') is not None:
                df.at[idx, 'altura_cm'] = bio['altura'] * 100
            altura_cm = df.at[idx, 'altura_cm']
            peso_kg = df.at[idx, 'peso_kg']
            if pd.notna(altura_cm) and pd.notna(peso_kg) and altura_cm > 0:
                df.at[idx, 'IMC'] = round(peso_kg / ((altura_cm/100)**2), 1)
            else:
                df.at[idx, 'IMC'] = np.nan
            if bio.get('pct_faulkner') is not None:
                df.at[idx, 'PctGordura_Faulkner'] = bio['pct_faulkner']
            if bio.get('pct_pollock3') is not None:
                df.at[idx, 'PctGordura_Pollock3'] = bio['pct_pollock3']
            if bio.get('pct_pollock7') is not None:
                df.at[idx, 'PctGordura_Pollock7'] = bio['pct_pollock7']
            pct_principal = bio.get('pct_pollock7') or bio.get('pct_pollock3') or bio.get('pct_faulkner')
            if pct_principal is not None:
                df.at[idx, 'Gordura_Corporal_%'] = pct_principal
            else:
                idade = df.at[idx, 'Idade']
                imc = df.at[idx, 'IMC']
                if pd.notna(imc) and pd.notna(idade):
                    df.at[idx, 'Gordura_Corporal_%'] = round(1.20*imc + 0.23*idade - 16.2, 1)
                else:
                    df.at[idx, 'Gordura_Corporal_%'] = np.nan
            peso = df.at[idx, 'peso_kg']
            gordura = df.at[idx, 'Gordura_Corporal_%']
            if pd.notna(peso) and pd.notna(gordura):
                massa_magra = round(peso * (1 - gordura/100), 1)
                df.at[idx, 'Massa_Magra_kg'] = massa_magra
                df.at[idx, 'Massa_Gorda_kg'] = round(peso - massa_magra, 1)
            else:
                df.at[idx, 'Massa_Magra_kg'] = np.nan
                df.at[idx, 'Massa_Gorda_kg'] = np.nan
                massa_magra = np.nan
            if bio.get('massa_muscular') is not None:
                df.at[idx, 'Massa_Muscular_Estimada_kg'] = round(bio['massa_muscular'], 1)
                df.at[idx, 'Massa_Muscular_Origem'] = 'Lee'
            else:
                if pd.notna(massa_magra):
                    df.at[idx, 'Massa_Muscular_Estimada_kg'] = round(massa_magra * 0.55, 1)
                    df.at[idx, 'Massa_Muscular_Origem'] = 'estimada'
                else:
                    df.at[idx, 'Massa_Muscular_Estimada_kg'] = np.nan
                    df.at[idx, 'Massa_Muscular_Origem'] = ''
            gordura_val = df.at[idx, 'Gordura_Corporal_%']
            idade_val = df.at[idx, 'Idade']
            if pd.notna(gordura_val) and pd.notna(idade_val):
                df.at[idx, 'Classificacao_Gordura'] = classif_gordura(gordura_val, idade_val)
            else:
                df.at[idx, 'Classificacao_Gordura'] = "Indefinido"
            imc_val = df.at[idx, 'IMC']
            imc_class = classif_imc(imc_val) if pd.notna(imc_val) else "Indefinido"
            gordura_class = df.at[idx, 'Classificacao_Gordura']
            df.at[idx, 'Estado_Fisico'] = estado_fisico(imc_class, gordura_class)
    return df

# =============================================
# CARTÕES (JSON) – COM LÓGICA DE FASES E RESET E ORDENAÇÃO
# =============================================
def carregar_cartoes_json(categoria):
    caminho = {
        "profissional": CAMINHO_CARTOES_PROFISSIONAL,
        "sub15": CAMINHO_CARTOES_SUB15,
        "sub17": CAMINHO_CARTOES_SUB17,
        "comissao_profissional": CAMINHO_CARTOES_COMISSAO_PROFISSIONAL,
        "comissao_sub15": CAMINHO_CARTOES_COMISSAO_SUB15,
        "comissao_sub17": CAMINHO_CARTOES_COMISSAO_SUB17,
    }.get(categoria)

    if not caminho or not os.path.exists(caminho):
        cartoes, datas = inicializar_cartoes_por_csvs(categoria, {})
        return cartoes, datas

    try:
        with open(caminho, 'r', encoding='utf-8') as f:
            dados = json.load(f)
            cartoes = dados.get('cartoes', {})
            cartoes = ordenar_historico_cartoes(cartoes)
            datas_globais = dados.get('datas_globais', {})
            return cartoes, datas_globais
    except Exception as e:
        print(f"Erro ao carregar JSON: {e}")
        return {}, {}

def salvar_cartoes_json(cartoes, categoria, datas_globais=None):
    if categoria == 'profissional':
        caminho = CAMINHO_CARTOES_PROFISSIONAL
    elif categoria == 'sub15':
        caminho = CAMINHO_CARTOES_SUB15
    elif categoria == 'sub17':
        caminho = CAMINHO_CARTOES_SUB17
    elif categoria == 'comissao_profissional':
        caminho = CAMINHO_CARTOES_COMISSAO_PROFISSIONAL
    elif categoria == 'comissao_sub15':
        caminho = CAMINHO_CARTOES_COMISSAO_SUB15
    elif categoria == 'comissao_sub17':
        caminho = CAMINHO_CARTOES_COMISSAO_SUB17
    else:
        caminho = os.path.join(DATA_DIR, f"cartoes_{categoria}.json")

    cartoes = ordenar_historico_cartoes(cartoes)

    dados = {'cartoes': cartoes}
    if datas_globais:
        dados_serializaveis = {}
        for id_jogador, lista_datas in datas_globais.items():
            if isinstance(lista_datas, (list, tuple)):
                dados_serializaveis[id_jogador] = [
                    d.strftime("%d/%m/%Y") if hasattr(d, 'strftime') else str(d)
                    for d in lista_datas
                ]
            else:
                dados_serializaveis[id_jogador] = str(lista_datas)
        dados['datas_globais'] = dados_serializaveis
    else:
        dados['datas_globais'] = {}

    with open(caminho, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

def jogador_suspenso(nome, cartoes):
    if nome not in cartoes:
        return False
    return cartoes[nome].get('suspenso_proxima', False)

# =============================================
# INICIALIZAR CARTÕES – COM RESET DE AMARELOS E VERIFICAÇÃO POR DATA
# =============================================
def inicializar_cartoes_por_df(df, categoria, canonico_para_ogol_id=None):
    if df.empty:
        return {}, {}

    df = df.sort_values('data_jogo', ascending=True)

    df_crono = carregar_cronograma(categoria.capitalize())
    if not df_crono.empty and 'data' in df_crono.columns:
        df_crono['data'] = pd.to_datetime(df_crono['data'], errors='coerce')
        datas_validas = df_crono['data'].dropna()
        datas_cronograma = sorted(datas_validas.dt.strftime('%Y-%m-%d').tolist())
    else:
        datas_cronograma = []

    cartoes = {}
    datas_globais = {}
    jogador_datas = {}

    for _, row in df.iterrows():
        nome = row.get('jogador')
        if pd.isna(nome) or str(nome).strip() == '':
            nome = row.get('nome_completo')
        if pd.isna(nome) or str(nome).strip() == '':
            continue
        nome = str(nome).strip()

        data_raw = row.get('data_jogo') or row.get('data')
        if data_raw:
            try:
                if isinstance(data_raw, pd.Timestamp):
                    data = data_raw.strftime('%Y-%m-%d')
                elif '/' in str(data_raw):
                    dt = datetime.strptime(str(data_raw).strip(), "%d/%m/%Y")
                    data = dt.strftime('%Y-%m-%d')
                else:
                    dt = pd.to_datetime(data_raw, dayfirst=True)
                    data = dt.strftime('%Y-%m-%d')
            except:
                try:
                    dt = pd.to_datetime(data_raw)
                    data = dt.strftime('%Y-%m-%d')
                except:
                    data = str(data_raw).strip()
            if nome not in jogador_datas:
                jogador_datas[nome] = set()
            jogador_datas[nome].add(data)

        if nome not in cartoes:
            id_ogol = row.get('id_ogol_jogador')
            if pd.isna(id_ogol):
                id_ogol = None
            else:
                try:
                    id_ogol = str(int(float(id_ogol))) if not pd.isna(id_ogol) else None
                except:
                    id_ogol = None

            cartoes[nome] = {
                'amarelos': 0,
                'vermelho': False,
                'suspenso_proxima': False,
                'historico': [],
                'id_ogol': id_ogol,
                'contador_amarelos_desde_reset': 0,
                'suspensoes_cumpridas': 0,
                'data_suspensao': None
            }

        adversario = row.get('adversario', 'N/I')
        competicao = row.get('Competicao', '')
        fase = row.get('Fase', '')
        amarelos = int(row.get('cartoes_amarelos', 0))
        vermelhos = int(row.get('cartoes_vermelhos', 0))

        # --- Amarelos ---
        if amarelos > 0:
            for _ in range(amarelos):
                cartoes[nome]['historico'].append({
                    'data': data,
                    'adversario': adversario,
                    'cor': 'amarelo',
                    'terceiro_amarelo': False,
                    'suspenso_causada': False,
                    'suspenso_cumprida': False,
                    'competicao': competicao,
                    'fase': fase
                })
                cartoes[nome]['contador_amarelos_desde_reset'] += 1
                if cartoes[nome]['contador_amarelos_desde_reset'] >= 3:
                    if cartoes[nome]['historico']:
                        ultimo = cartoes[nome]['historico'][-1]
                        ultimo['terceiro_amarelo'] = True
                        ultimo['suspenso_causada'] = True
                    cartoes[nome]['suspenso_proxima'] = True
                    cartoes[nome]['data_suspensao'] = data
                    cartoes[nome]['suspensoes_cumpridas'] = 0

        # --- Vermelhos ---
        if vermelhos > 0:
            for _ in range(vermelhos):
                cartoes[nome]['vermelho'] = True
                cartoes[nome]['historico'].append({
                    'data': data,
                    'adversario': adversario,
                    'cor': 'vermelho',
                    'terceiro_amarelo': False,
                    'suspenso_causada': True,
                    'suspenso_cumprida': False,
                    'competicao': competicao,
                    'fase': fase
                })
                cartoes[nome]['suspenso_proxima'] = True
                cartoes[nome]['contador_amarelos_desde_reset'] = 0
                cartoes[nome]['data_suspensao'] = data
                cartoes[nome]['suspensoes_cumpridas'] = 0

        cartoes[nome]['amarelos'] = cartoes[nome]['contador_amarelos_desde_reset']
        if cartoes[nome]['vermelho']:
            cartoes[nome]['suspenso_proxima'] = True

        id_ogol = cartoes[nome]['id_ogol']
        if id_ogol:
            if id_ogol not in datas_globais:
                datas_globais[id_ogol] = []
            if data not in datas_globais[id_ogol]:
                datas_globais[id_ogol].append(data)

    hoje = datetime.now(ZoneInfo("America/Sao_Paulo")).date()

    for nome, dados in cartoes.items():
        if not dados.get('suspenso_proxima', False):
            continue

        data_susp = dados.get('data_suspensao')
        if not data_susp:
            continue

        try:
            dt_susp = datetime.strptime(data_susp, "%Y-%m-%d").date()
        except:
            continue

        data_proximo_jogo = None
        competicao_susp = ''
        fase_susp = ''
        if datas_cronograma:
            datas_futuras = [d for d in datas_cronograma if d > data_susp]
            if datas_futuras:
                data_proximo_jogo_str = datas_futuras[0]
                data_proximo_jogo = datetime.strptime(data_proximo_jogo_str, "%Y-%m-%d").date()
                if df_crono is not None and not df_crono.empty:
                    jogos_prox = df_crono[df_crono['data'].dt.strftime('%Y-%m-%d') == data_proximo_jogo_str]
                    if not jogos_prox.empty:
                        competicao_susp = jogos_prox.iloc[0].get('competicao', '')
                        fase_susp = jogos_prox.iloc[0].get('fase', '')

        if data_proximo_jogo is None:
            data_proximo_jogo = dt_susp + timedelta(days=7)

        if data_proximo_jogo < hoje:
            data_str_prox = data_proximo_jogo.strftime("%Y-%m-%d")
            if data_str_prox not in jogador_datas.get(nome, set()):
                cartoes[nome]['historico'].append({
                    'data': data_str_prox,
                    'adversario': 'Suspensão cumprida (não jogou)',
                    'cor': 'suspensao_cumprida',
                    'terceiro_amarelo': False,
                    'suspenso_causada': False,
                    'suspenso_cumprida': True,
                    'competicao': competicao_susp,
                    'fase': fase_susp,
                    'observacao': f"Suspensão cumprida em {data_str_prox}"
                })
                cartoes[nome]['contador_amarelos_desde_reset'] = 0
                cartoes[nome]['amarelos'] = 0
                cartoes[nome]['suspenso_proxima'] = False
                cartoes[nome]['suspensoes_cumpridas'] = 0
                cartoes[nome]['data_suspensao'] = None

    cartoes = ordenar_historico_cartoes(cartoes)
    salvar_cartoes_json(cartoes, categoria, datas_globais)
    return cartoes, datas_globais

def inicializar_cartoes_por_csvs(categoria, canonico_para_ogol_id):
    df = carregar_estatisticas_partidas(categoria)
    return inicializar_cartoes_por_df(df, categoria, canonico_para_ogol_id)

# =============================================
# INICIALIZAR CARTÕES COMISSÃO
# =============================================
def inicializar_cartoes_comissao(categoria, df_comissao):
    st.info(f"🔄 Reinicializando cartões da comissão para {categoria}...")
    if categoria == 'comissao_profissional':
        pasta = PASTA_ESTATISTICAS_COMISSAO_PROFISSIONAL
    elif categoria == 'comissao_sub15':
        pasta = PASTA_ESTATISTICAS_COMISSAO_SUB15
    elif categoria == 'comissao_sub17':
        pasta = PASTA_ESTATISTICAS_COMISSAO_SUB17
    else:
        return {}, []
    if not pasta or not os.path.exists(pasta):
        st.warning(f"Pasta {pasta} não encontrada.")
        return {}, []
    lista_arquivos = [os.path.join(pasta, f) for f in os.listdir(pasta) if f.endswith('.csv') and f.startswith('jogo_')]
    if not lista_arquivos:
        st.warning("Nenhum CSV de estatísticas da comissão encontrado.")
        return {}, []

    arquivos_com_data = []
    for arq in lista_arquivos:
        data_jogo = extrair_data_jogo(arq)
        if not data_jogo:
            try:
                data_jogo = datetime.fromtimestamp(os.path.getmtime(arq))
            except:
                continue
        arquivos_com_data.append((data_jogo, arq))
    arquivos_com_data.sort(key=lambda x: x[0])
    datas_globais = [d.strftime("%Y-%m-%d") for d, _ in arquivos_com_data]

    cartoes = {}
    competicao_anterior = None
    fase_anterior = None
    ids_processados = set()

    df_crono = carregar_cronograma()
    mapa_jogo = {}
    if not df_crono.empty and 'id_jogo' in df_crono.columns:
        for _, row in df_crono.iterrows():
            jogo_id = row.get('id_jogo')
            if jogo_id:
                mapa_jogo[str(jogo_id)] = (row.get('competicao', 'Desconhecida'), row.get('fase', ''))

    for data_jogo, arq in arquivos_com_data:
        jogo_id = extrair_id_jogo(arq)
        if jogo_id is not None and jogo_id in ids_processados:
            continue
        if jogo_id is not None:
            ids_processados.add(jogo_id)

        try:
            df = pd.read_csv(arq, sep=';', encoding='utf-8-sig', on_bad_lines='skip')
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

            jogo_id_str = str(jogo_id) if jogo_id else ""
            if jogo_id_str in mapa_jogo:
                competicao_atual, fase_atual = mapa_jogo[jogo_id_str]
            else:
                competicao_atual = df['competicao'].iloc[0] if 'competicao' in df.columns else 'Desconhecida'
                fase_atual = df['fase'].iloc[0] if 'fase' in df.columns else ''
            adversario = df['adversario'].iloc[0] if 'adversario' in df.columns else 'Desconhecido'

            chave_fase_atual = f"{competicao_atual}_{fase_atual}"
            if competicao_anterior is not None and chave_fase_atual != f"{competicao_anterior}_{fase_anterior}":
                for dados in cartoes.values():
                    dados['amarelos'] = 0
                    dados['vermelho'] = False
                    dados['suspenso_proxima'] = False
            competicao_anterior = competicao_atual
            fase_anterior = fase_atual

            relacionados_nomes = set()
            for _, row in df.iterrows():
                nome = row.get('nome') or row.get('membro') or row.get('comissao') or row.get('staff')
                if pd.notna(nome):
                    canonico = mapear_nome_para_canonico(nome)
                    if canonico:
                        relacionados_nomes.add(canonico)

            for nome_membro, dados in list(cartoes.items()):
                if dados.get('suspenso_proxima', False):
                    nome_canonico = mapear_nome_para_canonico(nome_membro)
                    if nome_canonico not in relacionados_nomes:
                        dados['amarelos'] = 0
                        dados['vermelho'] = False
                        dados['suspenso_proxima'] = False
                        for ev in reversed(dados.get('historico', [])):
                            if ev.get('suspenso_causada') and not ev.get('suspenso_cumprida'):
                                ev['suspenso_cumprida'] = True
                                break

            suspensos_neste_jogo = set()
            for _, row in df.iterrows():
                nome_raw = row.get('nome') or row.get('membro') or row.get('comissao') or row.get('staff')
                if pd.isna(nome_raw):
                    continue
                nome = mapear_nome_para_canonico(nome_raw)
                if not nome:
                    continue

                amarelos = int(row.get('cartoes_amarelos', 0))
                vermelhos = int(row.get('cartoes_vermelhos', 0))
                if amarelos == 0 and vermelhos == 0:
                    continue

                if nome not in cartoes:
                    cartoes[nome] = {
                        'amarelos': 0,
                        'vermelho': False,
                        'suspenso_proxima': False,
                        'historico': []
                    }

                for _ in range(amarelos):
                    cartoes[nome]['amarelos'] += 1
                    terceiro = cartoes[nome]['amarelos'] >= 3
                    if terceiro:
                        cartoes[nome]['suspenso_proxima'] = True
                        suspensos_neste_jogo.add(nome)
                    cartoes[nome]['historico'].append({
                        'data': data_jogo.strftime("%d/%m/%Y"),
                        'adversario': adversario,
                        'competicao': competicao_atual,
                        'fase': fase_atual,
                        'cor': 'amarelo',
                        'terceiro_amarelo': terceiro,
                        'suspenso_causada': terceiro,
                        'suspenso_cumprida': False
                    })

                for _ in range(vermelhos):
                    cartoes[nome]['vermelho'] = True
                    cartoes[nome]['suspenso_proxima'] = True
                    suspensos_neste_jogo.add(nome)
                    cartoes[nome]['historico'].append({
                        'data': data_jogo.strftime("%d/%m/%Y"),
                        'adversario': adversario,
                        'competicao': competicao_atual,
                        'fase': fase_atual,
                        'cor': 'vermelho',
                        'terceiro_amarelo': False,
                        'suspenso_causada': True,
                        'suspenso_cumprida': False
                    })

        except Exception as e:
            st.warning(f"Erro ao processar {arq}: {e}")

    cartoes = ordenar_historico_cartoes(cartoes)
    salvar_cartoes_json(cartoes, categoria, datas_globais)
    st.success(f"✅ Cartões da comissão reinicializados para {categoria}.")
    return cartoes, datas_globais

# =============================================
# ESTATÍSTICAS DE PARTIDAS
# =============================================
def listar_arquivos_estatisticas(categoria="Profissional") -> List[str]:
    if categoria == "Profissional":
        pasta = PASTA_ESTATISTICAS_PROFISSIONAL
    elif categoria == "Sub-15":
        pasta = PASTA_ESTATISTICAS_SUB15
    elif categoria == "Sub-17":
        pasta = PASTA_ESTATISTICAS_SUB17
    else:
        pasta = PASTA_ESTATISTICAS_PROFISSIONAL
    if not os.path.exists(pasta):
        return []
    return [os.path.join(pasta, f) for f in os.listdir(pasta) if f.endswith('.csv') and f.startswith('jogo_')]

def carregar_estatisticas_partidas(categoria="Profissional") -> pd.DataFrame:
    arquivo = {
        "Profissional": "estatisticas_jogadores_profissional_2026.csv",
        "Sub-15": "estatisticas_jogadores_sub15_2026.csv",
        "Sub-17": "estatisticas_jogadores_sub17_2026.csv",
    }.get(categoria)
    if not arquivo:
        return pd.DataFrame()

    caminho = os.path.join(
        PASTA_ESTATISTICAS_PROFISSIONAL if categoria == "Profissional" else
        PASTA_ESTATISTICAS_SUB15 if categoria == "Sub-15" else
        PASTA_ESTATISTICAS_SUB17,
        arquivo
    )
    if not os.path.exists(caminho):
        fallback = os.path.join(DATA_DIR, arquivo)
        if os.path.exists(fallback):
            caminho = fallback
        else:
            script_dir = Path(__file__).resolve().parent
            abs_caminho = script_dir / "data" / "estatisticas_jogadores" / arquivo
            if abs_caminho.exists():
                caminho = str(abs_caminho)
            else:
                return pd.DataFrame()

    try:
        df = pd.read_csv(caminho, sep=';', encoding='utf-8-sig')
        if 'data_jogo' in df.columns:
            df['data_jogo'] = pd.to_datetime(df['data_jogo'], dayfirst=True, errors='coerce')
        for col in ['fase', 'competicao', 'adversario', 'jogador']:
            if col in df.columns:
                df[col] = df[col].astype(str)
        return sanitizar_dataframe(df)
    except Exception as e:
        print(f"Erro ao carregar {caminho}: {e}")
        return pd.DataFrame()

def precomputar_scores_posicionais(df, df_stats_partidas):
    if df_stats_partidas.empty:
        for col in ['starts', 'jogos_90min', 'minutos_totais_partidas']:
            if col not in df.columns:
                df[col] = 0
        return df

    colunas_necessarias = ['starts', 'jogos_90min', 'minutos_totais']
    colunas_presentes = [col for col in colunas_necessarias if col in df_stats_partidas.columns]
    if not colunas_presentes:
        for col in ['starts', 'jogos_90min', 'minutos_totais_partidas']:
            if col not in df.columns:
                df[col] = 0
        return df

    df_merged = df.copy()

    if 'apelido' in df_merged.columns:
        df_merged['nome_canonico'] = df_merged['apelido'].apply(mapear_nome_para_canonico)
    else:
        df_merged['nome_canonico'] = None

    df_stats = df_stats_partidas.copy()
    if 'minutos_totais' in df_stats.columns:
        df_stats = df_stats.rename(columns={'minutos_totais': 'minutos_totais_partidas'})

    if 'jogador_canonico' not in df_stats.columns:
        if 'jogador' in df_stats.columns:
            df_stats['jogador_canonico'] = df_stats['jogador'].apply(mapear_nome_para_canonico)
        elif 'nome_completo' in df_stats.columns:
            df_stats['jogador_canonico'] = df_stats['nome_completo'].apply(mapear_nome_para_canonico)
        else:
            for col in ['starts', 'jogos_90min', 'minutos_totais_partidas']:
                if col not in df_merged.columns:
                    df_merged[col] = 0
            return df_merged

    colunas_merge = ['jogador_canonico', 'starts', 'jogos_90min']
    if 'minutos_totais_partidas' in df_stats.columns:
        colunas_merge.append('minutos_totais_partidas')
    df_stats = df_stats[colunas_merge]
    df_stats = df_stats.dropna(subset=['jogador_canonico'])

    df_merged = df_merged.merge(df_stats, left_on='nome_canonico', right_on='jogador_canonico', how='left')
    df_merged.drop(columns=['jogador_canonico', 'nome_canonico'], errors='ignore', inplace=True)

    for col in ['starts', 'jogos_90min', 'minutos_totais_partidas']:
        if col in df_merged.columns:
            df_merged[col] = df_merged[col].fillna(0).astype(int)
        else:
            df_merged[col] = 0

    return sanitizar_dataframe(df_merged)

# =============================================
# FUNÇÕES DE ESCALAÇÃO E FORMAÇÃO (CORRIGIDO)
# =============================================
def interpretar_formacao(formacao_str):
    """
    Interpreta uma string de formação (ex: '4-4-2') e retorna a lista de posições.
    Retorna (None, None, None, None) se a formação for inválida ou vazia.
    """
    if not formacao_str or not isinstance(formacao_str, str):
        return None, None, None, None
    partes = formacao_str.split('-')
    if len(partes) < 3:
        return None, None, None, None
    try:
        nums = [int(p) for p in partes]
        if sum(nums) != 10:
            return None, None, None, None
        defensores = nums[0]
        atacantes = nums[-1]
        meio_campistas = sum(nums[1:-1])
        posicoes = [('Goleiro', 'Goleiro')]
        for i in range(defensores):
            posicoes.append((f'Defensor {i+1}', 'Defensor'))
        for i in range(meio_campistas):
            posicoes.append((f'Meio-Campista {i+1}', 'Meio-Campo'))
        for i in range(atacantes):
            posicoes.append((f'Atacante {i+1}', 'Atacante'))
        return defensores, meio_campistas, atacantes, posicoes
    except:
        return None, None, None, None

def obter_jogadores_para_posicao(df, pos_tipo, excluidos, cartoes, incluir_lesionados=False):
    candidatos = df.copy()
    if pos_tipo == 'Goleiro':
        candidatos = candidatos[candidatos['Posicao_Principal'] == 'Goleiro']
    else:
        candidatos = candidatos[~candidatos['Posicao_Principal'].isin(['Goleiro'])]
    candidatos = candidatos[~candidatos['nome_completo'].isin(excluidos)]
    candidatos = candidatos[~candidatos['nome_completo'].apply(lambda x: jogador_suspenso(mapear_nome_para_canonico(x), cartoes))]
    if not incluir_lesionados and 'lesionado' in candidatos.columns:
        candidatos = candidatos[~candidatos['lesionado']]
    return candidatos

def obter_atributos_chave(posicao):
    mapa = {
        'Goleiro': ['reflexos','defesas_goleiro','jogo_aereo_goleiro','comando_area','saida_gol'],
        'Zagueiro': ['desarme','marcacao','cabecada','forca_fisica','antecipacao','posicionamento'],
        'Lateral': ['cruzamentos','drible','passe','velocidade_maxima','aceleracao','resistencia'],
        'Lateral Direito': ['cruzamentos','drible','passe','velocidade_maxima','aceleracao','resistencia'],
        'Lateral Esquerdo': ['cruzamentos','drible','passe','velocidade_maxima','aceleracao','resistencia'],
        'Volante': ['desarme','marcacao','passe','posicionamento','intensidade_trabalho','visao_jogo'],
        'Meia': ['visao_jogo','passe','criatividade','tecnica','primeiro_controle','chutes_longe'],
        'Meia-Central': ['visao_jogo','passe','criatividade','tecnica','desarme','posicionamento'],
        'Meia-Atacante': ['visao_jogo','passe','criatividade','tecnica','primeiro_controle','finalizacao'],
        'Ponta': ['drible','aceleracao','velocidade_maxima','cruzamentos','tecnica','finalizacao'],
        'Ponta Direita': ['drible','aceleracao','velocidade_maxima','cruzamentos','tecnica','finalizacao'],
        'Ponta Esquerda': ['drible','aceleracao','velocidade_maxima','cruzamentos','tecnica','finalizacao'],
        'Centroavante': ['finalizacao','cabecada','primeiro_controle','composicao','forca_fisica','movimentacao_sem_bola'],
        'Segundo Atacante': ['movimentacao_sem_bola','primeiro_controle','tecnica','visao_jogo','passe','drible'],
        'Defensor': ['desarme','marcacao','cabecada','forca_fisica','antecipacao','posicionamento'],
        'Meio-Campo': ['passe','visao_jogo','criatividade','desarme','intensidade_trabalho'],
    }
    return mapa.get(posicao, ['Rating_Geral_FM26'])

# =============================================
# AUTENTICAÇÃO DE USUÁRIOS
# =============================================
ARQUIVO_USUARIOS = "usuarios.json"

def carregar_usuarios():
    if not os.path.exists(ARQUIVO_USUARIOS):
        senha_admin = "@W.d06302005"
        hash_admin = bcrypt.hashpw(senha_admin.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        usuarios = {"Guibfpinto": hash_admin}
        with open(ARQUIVO_USUARIOS, 'w', encoding='utf-8') as f:
            json.dump(usuarios, f, indent=2, ensure_ascii=False)
        return usuarios
    try:
        with open(ARQUIVO_USUARIOS, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return {"Guibfpinto": ""}

def autenticar_usuario(usuario, senha):
    usuarios = carregar_usuarios()
    if usuario not in usuarios:
        return False
    hash_senha = usuarios[usuario].encode('utf-8')
    return bcrypt.checkpw(senha.encode('utf-8'), hash_senha)

def listar_usuarios():
    return list(carregar_usuarios().keys())

def adicionar_usuario(usuario, senha):
    usuarios = carregar_usuarios()
    if usuario in usuarios:
        return False
    hash_novo = bcrypt.hashpw(senha.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    usuarios[usuario] = hash_novo
    with open(ARQUIVO_USUARIOS, 'w', encoding='utf-8') as f:
        json.dump(usuarios, f, indent=2, ensure_ascii=False)
    return True

def remover_usuario(usuario):
    if usuario == "Guibfpinto":
        return False
    usuarios = carregar_usuarios()
    if usuario in usuarios:
        del usuarios[usuario]
        with open(ARQUIVO_USUARIOS, 'w', encoding='utf-8') as f:
            json.dump(usuarios, f, indent=2, ensure_ascii=False)
        return True
    return False

# =============================================
# RELATÓRIOS E EXPORTAÇÃO
# =============================================
def gerar_relatorio_completo_texto(df, nome_categoria):
    if df is None or df.empty:
        return f"Sem dados para {nome_categoria}"
    texto = f"Relatório - {nome_categoria}\n\nTotal: {len(df)}\n"
    if 'Idade' in df.columns:
        texto += f"Idade média: {df['Idade'].mean():.1f}\n"
    if 'Rating_Geral_FM26' in df.columns:
        texto += f"Rating médio: {df['Rating_Geral_FM26'].mean():.1f}\n"
    return texto

def exportar_para_excel(df, nome_categoria, caminho_arquivo):
    try:
        df.to_excel(caminho_arquivo, index=False, engine='openpyxl')
        return True
    except Exception as e:
        print(f"Erro na exportação: {e}")
        return False

def exportar_para_powerbi(df, nome_categoria, caminho_json_fotos, pasta_fotos, caminho_arquivo):
    return exportar_para_excel(df, nome_categoria, caminho_arquivo)

# =============================================
# CORREÇÕES DE NOMES DE TIMES
# =============================================
CORRECOES_NOMES_TIMES = {
    "Serra Talhada": "Serra-ES",
    "Sport Brasil ES": "Sport Brasil Capixaba",
}

def corrigir_nome_time(nome):
    return CORRECOES_NOMES_TIMES.get(nome, nome)

# =============================================
# FUNÇÕES DA FASTAPI (USANDO BANCO LOCAL)
# =============================================
def _chamar_api(endpoint, params=None, tentativa=1):
    return None

def verificar_jogo_ao_vivo(team_id=None):
    conn = sqlite3.connect('meu_futebol.db')
    cursor = conn.cursor()
    query = """
        SELECT id FROM jogos
        WHERE status IN ('1H', '2H', 'HT', 'ET', 'P')
    """
    params = []
    if team_id:
        query += " AND (time_casa_id = ? OR time_fora_id = ?)"
        params.extend([team_id, team_id])
    query += " LIMIT 1"
    cursor.execute(query, params)
    row = cursor.fetchone()
    conn.close()
    if row:
        return row[0]
    return None

def obter_detalhes_jogo(fixture_id):
    conn = sqlite3.connect('meu_futebol.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("""
        SELECT j.*, tc.nome AS time_casa_nome, tf.nome AS time_fora_nome, v.nome AS estadio
        FROM jogos j
        LEFT JOIN times tc ON j.time_casa_id = tc.id
        LEFT JOIN times tf ON j.time_fora_id = tf.id
        LEFT JOIN venues v ON j.venue_id = v.id
        WHERE j.id = ?
    """, (fixture_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return None
    dados = dict(row)
    return {
        "fixture": {
            "id": dados["id"],
            "date": dados.get("data_hora", ""),
            "status": {"short": dados.get("status", "NS")},
            "venue": {"name": dados.get("estadio", "Estádio")},
            "referee": dados.get("arbitro", "Não informado")
        },
        "teams": {
            "home": {"id": dados["time_casa_id"], "name": dados.get("time_casa_nome", "Casa")},
            "away": {"id": dados["time_fora_id"], "name": dados.get("time_fora_nome", "Fora")}
        },
        "goals": {
            "home": dados.get("gols_casa", 0),
            "away": dados.get("gols_fora", 0)
        }
    }

def obter_eventos_jogo(fixture_id):
    conn = sqlite3.connect('meu_futebol.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("""
        SELECT e.*, el.nome AS jogador_nome, el.apelido AS jogador_apelido
        FROM eventos e
        LEFT JOIN elenco el ON e.jogador_id = el.id
        WHERE e.jogo_id = ?
        ORDER BY e.tempo ASC
    """, (fixture_id,))
    rows = cursor.fetchall()
    conn.close()
    eventos = []
    for r in rows:
        d = dict(r)
        eventos.append({
            "time": {"elapsed": d.get("tempo", 0)},
            "type": d.get("tipo", ""),
            "detail": d.get("detalhes", ""),
            "player": {"name": d.get("jogador_nome") or d.get("jogador_apelido") or "Desconhecido"},
            "team": {"name": ""}
        })
    return eventos

def obter_estatisticas_jogo(fixture_id):
    return []

def obter_lineups_completos(fixture_id):
    return []

def obter_players_stats(fixture_id):
    return []

def buscar_jogos_por_competicao(league_id, season_id, team_id, data_inicio, data_fim):
    conn = sqlite3.connect('meu_futebol.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    query = """
        SELECT * FROM jogos
        WHERE (time_casa_id = ? OR time_fora_id = ?)
        AND data_hora BETWEEN ? AND ?
        ORDER BY data_hora
    """
    cursor.execute(query, (team_id, team_id, data_inicio, data_fim))
    rows = cursor.fetchall()
    conn.close()
    jogos = []
    for row in rows:
        d = dict(row)
        jogos.append({
            'id': d['id'],
            'data': d['data_hora'],
            'status_code': d['status'],
            'time_casa': d['time_casa_nome'] if 'time_casa_nome' in d else '',
            'time_fora': d['time_fora_nome'] if 'time_fora_nome' in d else '',
            'gols_casa': d['gols_casa'],
            'gols_fora': d['gols_fora']
        })
    return jogos

def gerar_relatorio_excel(fixture_id, time_casa_titulares=None, time_casa_reservas=None):
    print("📊 Gerando relatório Excel a partir do banco local...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório"
        ws['A1'] = f"Relatório da Partida {fixture_id}"
        pasta = RELATORIOS_DIR
        os.makedirs(pasta, exist_ok=True)
        caminho = os.path.join(pasta, f"relatorio_{fixture_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(caminho)
        print(f"✅ Relatório salvo em {caminho}")
    except Exception as e:
        print(f"⚠️ Erro ao gerar relatório: {e}")

# =============================================
# FORMATAR PLANILHA EXCEL
# =============================================
def formatar_planilha(ws, titulo):
    ws.title = titulo
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        if col:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = adjusted_width

def obter_historico_clubes(jogador_row):
    historico = jogador_row.get('historico')
    if pd.isna(historico) or not historico:
        return "Nenhum histórico de clubes registrado."
    return str(historico).strip()

# =============================================
# FORMATAR CARTÕES PARA EXIBIÇÃO (COM RESET)
# =============================================
def formatar_cartoes(cartoes: dict, nome_jogador: str = None) -> str:
    if not cartoes:
        return "Nenhum dado de cartões disponível."

    if nome_jogador:
        dados = None
        for chave, valor in cartoes.items():
            if chave.lower() == nome_jogador.lower() or mapear_nome_para_canonico(chave) == nome_jogador:
                dados = valor
                break
            if chave == nome_jogador:
                dados = valor
                break
        if not dados:
            return f"Jogador '{nome_jogador}' não encontrado."

        amarelos_atuais = dados.get('amarelos', 0)
        vermelho = "Sim" if dados.get('vermelho', False) else "Não"
        suspenso = "Sim" if dados.get('suspenso_proxima', False) else "Não"
        historico = dados.get('historico', [])
        ja_cumpriu = any(ev.get('cor') == 'suspensao_cumprida' for ev in historico)
        cumprida = "Sim" if ja_cumpriu else "Não"

        linhas = [
            f"📋 **Cartões de {nome_jogador}**",
            f"  🟨 Amarelos atuais (não pagos): {amarelos_atuais}",
            f"  🟥 Vermelho direto (ativo): {vermelho}",
            f"  ⚠️ Suspenso para o próximo jogo: {suspenso}",
            f"  ✅ Suspensão cumprida? {cumprida}",
            ""
        ]
        if historico:
            historico_ordenado = sorted(historico, key=lambda x: x['data'], reverse=True)
            linhas.append("  **Histórico (últimos eventos):**")
            for ev in historico_ordenado[-8:]:
                data = ev.get('data', 'data desconhecida')
                adv = ev.get('adversario', 'adversário')
                cor = ev.get('cor', '')
                cumprida_evento = "Sim" if ev.get('suspenso_cumprida', False) else "Não"
                if cor == 'amarelo':
                    emoji = "🟨"
                    extra = " (3º amarelo!)" if ev.get('terceiro_amarelo', False) else ""
                elif cor == 'vermelho':
                    emoji = "🟥"
                    extra = " (expulsão)"
                elif cor == 'suspensao_cumprida':
                    emoji = "🔄"
                    extra = " (suspensão cumprida)"
                else:
                    emoji = "ℹ️"
                    extra = ""
                linha = f"    {emoji} {data} vs {adv} - {cor}{extra} [Cumprida: {cumprida_evento}]"
                if ev.get('competicao'):
                    linha += f" [{ev.get('competicao')}"
                    if ev.get('fase'):
                        linha += f" - {ev.get('fase')}"
                    linha += "]"
                linhas.append(linha)
        else:
            linhas.append("  Nenhum cartão registrado.")
        return "\n".join(linhas)

    else:
        linhas = ["📊 **RESUMO DE CARTÕES DE TODOS OS JOGADORES**", ""]
        linhas.append(f"{'Jogador':<25} {'Atuais':<10} {'Vermelho':<10} {'Suspenso':<10} {'Cumprida':<15}")
        linhas.append("-" * 70)
        for jog, dados in sorted(cartoes.items()):
            amarelos_atuais = dados.get('amarelos', 0)
            vermelho = "Sim" if dados.get('vermelho', False) else "Não"
            suspenso = "Sim" if dados.get('suspenso_proxima', False) else "Não"
            historico = dados.get('historico', [])
            ja_cumpriu = any(ev.get('cor') == 'suspensao_cumprida' for ev in historico)
            cumprida = "Sim" if ja_cumpriu else "Não"
            linhas.append(f"{jog:<25} {amarelos_atuais:<10} {vermelho:<10} {suspenso:<10} {cumprida:<15}")
        return "\n".join(linhas)

# =============================================
# CONTROLE DE REINICIALIZAÇÃO AUTOMÁTICA (DIÁRIA)
# =============================================
ARQUIVO_CONTROLE_REINICIALIZACAO = os.path.join(DATA_DIR, "ultima_reinicializacao.json")

def _obter_data_brt_atual() -> str:
    agora = datetime.now(ZoneInfo("America/Sao_Paulo"))
    return agora.strftime("%Y-%m-%d")

def _carregar_ultima_reinicializacao() -> Optional[str]:
    if not os.path.exists(ARQUIVO_CONTROLE_REINICIALIZACAO):
        return None
    try:
        with open(ARQUIVO_CONTROLE_REINICIALIZACAO, 'r', encoding='utf-8') as f:
            dados = json.load(f)
            return dados.get('data', None)
    except:
        return None

def _salvar_ultima_reinicializacao(data: str):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(ARQUIVO_CONTROLE_REINICIALIZACAO, 'w', encoding='utf-8') as f:
        json.dump({'data': data}, f, ensure_ascii=False, indent=2)

def verificar_e_reinicializar_cartoes(categoria: str) -> bool:
    data_atual = _obter_data_brt_atual()
    ultima_data = _carregar_ultima_reinicializacao()
    if ultima_data is None or ultima_data != data_atual:
        chave_categoria = {
            "Profissional": "profissional",
            "Sub-15": "sub15",
            "Sub-17": "sub17",
        }.get(categoria, categoria.lower())
        novos_cartoes, _ = inicializar_cartoes_por_csvs(chave_categoria, {})
        _salvar_ultima_reinicializacao(data_atual)
        return True
    return False

# =============================================
# STUBS PARA FUNÇÕES USADAS EM PÁGINAS
# =============================================
def exportar_escalacao_excel(df, nome_arquivo="escalacao.xlsx"):
    try:
        df.to_excel(nome_arquivo, index=False, engine='openpyxl')
        return True
    except Exception as e:
        print(f"Erro ao exportar escalação: {e}")
        return False

def exportar_escalacao_pdf(df, nome_arquivo="escalacao.pdf"):
    try:
        return True
    except Exception as e:
        print(f"Erro ao exportar PDF: {e}")
        return False

def gerar_relatorio_diretoria(df, categoria):
    if df is None or df.empty:
        return f"Sem dados para gerar relatório da diretoria – {categoria}"
    texto = f"RELATÓRIO PARA DIRETORIA – {categoria}\n"
    texto += f"Total de membros: {len(df)}\n"
    if 'nome' in df.columns:
        texto += f"Principais: {', '.join(df['nome'].head(5).tolist())}\n"
    return texto

def gerar_relatorio_jogador(row, categoria):
    if row is None:
        return "Jogador não encontrado."
    nome = row.get('nome_completo', 'N/I')
    apelido = row.get('apelido', 'N/I')
    posicao = row.get('Posicao_Principal', 'N/I')
    idade = row.get('Idade', 'N/I')
    rating = row.get('Rating_Geral_FM26', 'N/I')
    texto = f"RELATÓRIO DO JOGADOR – {categoria}\n"
    texto += f"Nome: {nome}\nApelido: {apelido}\nPosição: {posicao}\nIdade: {idade}\nRating FM26: {rating}\n"
    return texto

def gerar_relatorio_comissao(df, categoria):
    if df is None or df.empty:
        return f"Sem dados para gerar relatório da comissão – {categoria}"
    texto = f"RELATÓRIO DA COMISSÃO TÉCNICA – {categoria}\n"
    texto += f"Total de membros: {len(df)}\n"
    if 'nome' in df.columns:
        texto += f"Membros: {', '.join(df['nome'].tolist())}\n"
    if 'cargo' in df.columns:
        cargos = df['cargo'].value_counts().to_dict()
        texto += "Distribuição por cargo:\n"
        for cargo, qtd in cargos.items():
            texto += f"  {cargo}: {qtd}\n"
    return texto